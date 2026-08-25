from io import BytesIO

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook as openpyxl_load_workbook


st.set_page_config(
    page_title="Run Rate Analysis",
    page_icon="◆",
    layout="wide",
    initial_sidebar_state="expanded",
)

GOLD = "#D4AF37"
TEXT = "#F3F0E7"
MUTED = "#A8A397"
CARD = "#111111"
REVENUE_RATE = 0.006  # Revenue = Net Sales × 0.60 / 100

MONTHS = {
    1: "July",
    2: "August",
    3: "September",
    4: "October",
    5: "November",
    6: "December",
    7: "January",
    8: "February",
    9: "March",
}

REQUIRED = {
    "Emp Code",
    "Employee Name",
    "FY 26 TGT EQ NS",
    "Equity NS Ach YTD June",
}


APP_PAGE = st.sidebar.radio(
    "Page",
    [
        "1 · Run Rate Analysis",
        "2 · FINAL Sheet Analysis",
    ],
)


st.markdown(
    """
<style>
.stApp {
    background:#070707;
    color:#F3F0E7
}

[data-testid="stSidebar"] {
    background:#0B0B0B;
    border-right:1px solid rgba(212,175,55,.24)
}

[data-testid="stSidebar"] * {
    color:#F3F0E7
}

.hero {
    border:1px solid rgba(212,175,55,.25);
    border-radius:22px;
    padding:24px 26px;
    margin-bottom:18px;
    background:linear-gradient(
        110deg,
        rgba(212,175,55,.10),
        rgba(255,255,255,.015)
    )
}

.eyebrow {
    color:#D4AF37;
    font-size:.76rem;
    letter-spacing:.17em;
    font-weight:750;
    text-transform:uppercase
}

.hero-title {
    font-size:clamp(2rem,3vw,3.1rem);
    line-height:1.05;
    font-weight:780;
    margin-top:8px
}

.hero-sub {
    color:#A8A397;
    margin-top:10px;
    font-size:.96rem;
    line-height:1.6;
    max-width:1100px
}

.section-title {
    font-size:1.28rem;
    font-weight:750;
    margin-top:18px
}

.section-note {
    color:#A8A397;
    font-size:.86rem;
    margin:4px 0 13px;
    line-height:1.5
}

.kpi {
    border:1px solid rgba(212,175,55,.24);
    border-radius:17px;
    padding:16px;
    min-height:120px;
    background:linear-gradient(145deg,#121212,#0D0D0D)
}

.kpi-label {
    color:#A8A397;
    font-size:.72rem;
    letter-spacing:.04em;
    text-transform:uppercase;
    font-weight:650
}

.kpi-value {
    font-size:1.62rem;
    font-weight:780;
    margin-top:7px;
    color:#D4AF37
}

.kpi-foot {
    color:#A8A397;
    font-size:.72rem;
    margin-top:7px;
    line-height:1.4
}

.callout {
    background:linear-gradient(
        145deg,
        rgba(212,175,55,.08),
        rgba(255,255,255,.01)
    );
    border:1px solid rgba(212,175,55,.24);
    border-left:3px solid #D4AF37;
    border-radius:14px;
    padding:15px 17px;
    color:#CBC6BA;
    font-size:.88rem;
    margin:8px 0 18px;
    line-height:1.65
}

.management {
    border:1px solid rgba(212,175,55,.24);
    border-radius:16px;
    padding:17px 18px;
    background:rgba(212,175,55,.055);
    color:#D7D2C7;
    line-height:1.7;
    min-height:175px
}

.management b {
    color:#D4AF37
}

[data-testid="stDataFrame"] {
    border:1px solid rgba(212,175,55,.20);
    border-radius:14px;
    overflow:hidden
}
</style>
""",
    unsafe_allow_html=True,
)


def clean_text(value):
    if value is None or (
        isinstance(value, float) and np.isnan(value)
    ):
        return ""

    return " ".join(str(value).strip().split())


def unique_headers(values):
    seen = {}
    output = []

    for index, value in enumerate(values):
        base = clean_text(value) or f"Unnamed_{index}"
        count = seen.get(base, 0)

        output.append(
            base if count == 0 else f"{base}.{count}"
        )

        seen[base] = count + 1

    return output


@st.cache_data(show_spinner=False)
def load_workbook(file_bytes):
    excel = pd.ExcelFile(
        BytesIO(file_bytes),
        engine="openpyxl",
    )

    source_sheet = (
        "RM Retail Sales"
        if "RM Retail Sales" in excel.sheet_names
        else excel.sheet_names[0]
    )

    raw = pd.read_excel(
        excel,
        sheet_name=source_sheet,
        header=None,
    )

    header_row = None

    for index in range(min(30, len(raw))):
        values = {
            clean_text(value)
            for value in raw.iloc[index].tolist()
        }

        if (
            "FY 26 TGT EQ NS" in values
            and "Equity NS Ach YTD June" in values
        ):
            header_row = index
            break

    if header_row is None:
        raise ValueError(
            "Could not locate the RM header row "
            "in the uploaded workbook."
        )

    df = raw.iloc[header_row + 1:].copy()
    df.columns = unique_headers(
        raw.iloc[header_row].tolist()
    )
    df = df.dropna(how="all").copy()

    missing = sorted(REQUIRED - set(df.columns))

    if missing:
        raise ValueError(
            "Missing required columns: "
            + ", ".join(missing)
        )

    for column in [
        "FY 26 TGT EQ NS",
        "Equity NS Ach YTD June",
    ]:
        df[column] = pd.to_numeric(
            df[column],
            errors="coerce",
        )

    for column in [
        "Emp Code",
        "Employee Name",
        "Status",
        "Type",
        "ZONE",
        "REGION",
    ]:
        if column in df.columns:
            df[column] = (
                df[column]
                .fillna("Unknown")
                .astype(str)
                .str.strip()
            )

    df["Emp Code"] = df["Emp Code"].str.replace(
        r"\.0$",
        "",
        regex=True,
    )

    if "MKT TYPE.1" in df.columns:
        market_source = "MKT TYPE.1"
    elif "MKT TYPE" in df.columns:
        market_source = "MKT TYPE"
    else:
        market_source = None

    if market_source:
        df["Market Type"] = (
            df[market_source]
            .fillna("Unknown")
            .astype(str)
            .str.strip()
        )
    else:
        df["Market Type"] = "Unknown"

    market_normalization = {
        "B30-SELECT": "B30",
        "B30 SELECT": "B30",
        "B30 SELECTED": "B30",
        "T30-EXT": "T30",
        "T30 EXT": "T30",
        "T30 EXTENDED": "T30",
    }

    normalized = (
        df["Market Type"]
        .str.upper()
        .str.strip()
    )

    df["Market Type"] = normalized.replace(
        market_normalization
    )

    df.loc[
        df["Market Type"].eq(""),
        "Market Type",
    ] = "Unknown"

    return df


def project(df, months, uplift):
    result = df.copy()

    result["Current Monthly RR"] = (
        result["Equity NS Ach YTD June"] / 3.0
    )

    result["Scenario Monthly RR"] = (
        result["Current Monthly RR"]
        * (1 + uplift / 100)
    )

    result["Projected Future NS"] = (
        result["Scenario Monthly RR"] * months
    )

    result["Projected Final NS"] = (
        result["Equity NS Ach YTD June"]
        + result["Projected Future NS"]
    )

    result["Projected Achievement %"] = np.where(
        result["FY 26 TGT EQ NS"] > 0,
        (
            result["Projected Final NS"]
            / result["FY 26 TGT EQ NS"]
            * 100
        ),
        np.nan,
    )

    result["Qualifies"] = (
        result["Projected Achievement %"] >= 100
    )

    return result


def summarize(frame, label, uplift, baseline):
    target = frame["FY 26 TGT EQ NS"].sum()
    projected = frame["Projected Final NS"].sum()

    qualifying = (
        frame["Projected Achievement %"] >= 100
    )

    baseline_qualifying = (
        baseline["Projected Achievement %"] >= 100
    )

    qualifying_ns = frame.loc[
        qualifying,
        "Projected Final NS",
    ].sum()

    baseline_projected = (
        baseline["Projected Final NS"].sum()
    )

    return {
        "Scenario": label,
        "RR Uplift %": uplift,
        "Total RMs": len(frame),
        "Target": target,
        "Projected NS": projected,
        "Portfolio Achievement %": (
            projected / target * 100
            if target else 0
        ),
        "RMs ≥100%": int(qualifying.sum()),
        "Qualification Rate %": (
            qualifying.mean() * 100
            if len(frame) else 0
        ),
        "New Qualifiers": int(
            (
                qualifying
                & ~baseline_qualifying
            ).sum()
        ),
        "Qualifying RM NS": qualifying_ns,
        "Qualifying NS Contribution %": (
            qualifying_ns / projected * 100
            if projected else 0
        ),
        "Revenue": projected * REVENUE_RATE,
        "Incremental NS": (
            projected - baseline_projected
        ),
        "Incremental Revenue": (
            projected - baseline_projected
        ) * REVENUE_RATE,
    }


def fmt(value):
    if value is None or pd.isna(value):
        return "—"

    return f"{float(value):,.2f}"


def pct(value):
    if value is None or pd.isna(value):
        return "—"

    return f"{float(value):,.1f}%"


def section(title, note=""):
    st.html(
        f"""
        <div class="section-title">{title}</div>
        <div class="section-note">{note}</div>
        """
    )


def kpi(label, value, foot=""):
    st.html(
        f"""
        <div class="kpi">
            <div class="kpi-label">{label}</div>
            <div class="kpi-value">{value}</div>
            <div class="kpi-foot">{foot}</div>
        </div>
        """
    )


def show_table(df, height=None):
    display = df.copy()

    numeric = display.select_dtypes(
        include=np.number
    ).columns

    display[numeric] = display[numeric].round(2)

    args = {
        "data": display,
        "width": "stretch",
        "hide_index": True,
    }

    if height is not None:
        args["height"] = height

    st.dataframe(**args)


def chart_style(fig, height=390):
    fig.update_layout(
        template=None,
        height=height,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor=CARD,
        font=dict(color=TEXT),
        margin=dict(
            l=45,
            r=30,
            t=65,
            b=45,
        ),
        xaxis=dict(
            gridcolor="rgba(212,175,55,.10)",
            zeroline=False,
        ),
        yaxis=dict(
            gridcolor="rgba(212,175,55,.10)",
            zeroline=False,
        ),
        hoverlabel=dict(
            bgcolor="#111111",
            font_color=TEXT,
        ),
    )

    return fig


@st.cache_data(show_spinner=False)
def load_final_analysis(file_bytes):
    """
    Read calculated values and formulas from
    the authoritative FINAL worksheet.
    """

    formula_book = openpyxl_load_workbook(
        BytesIO(file_bytes),
        data_only=False,
        read_only=False,
    )

    value_book = openpyxl_load_workbook(
        BytesIO(file_bytes),
        data_only=True,
        read_only=False,
    )

    if "FINAL" not in formula_book.sheetnames:
        raise ValueError(
            "The uploaded workbook does not "
            "contain a FINAL sheet."
        )

    formula_sheet = formula_book["FINAL"]
    value_sheet = value_book["FINAL"]

    def value(cell):
        return value_sheet[cell].value

    def formula(cell):
        raw = formula_sheet[cell].value

        if hasattr(raw, "text"):
            return raw.text

        if (
            isinstance(raw, str)
            and raw.startswith("=")
        ):
            return raw

        return "Input"

    def rows(specification):
        return pd.DataFrame(
            [
                {
                    heading: value(cell)
                    for heading, cell in row.items()
                }
                for row in specification
            ]
        )

    result = {}

    result["gross_sales_summary"] = rows(
        [
            {
                "Asset Class": "B6",
                "Target": "C6",
                "YTD": "D6",
                "Achievement %": "E6",
            },
            {
                "Asset Class": "B7",
                "Target": "C7",
                "YTD": "D7",
                "Achievement %": "E7",
            },
            {
                "Asset Class": "B8",
                "Target": "C8",
                "YTD": "D8",
                "Achievement %": "E8",
            },
            {
                "Asset Class": "B9",
                "Target": "C9",
                "YTD": "D9",
                "Achievement %": "E9",
            },
        ]
    )

    def run_rate_table(label_cells, start_columns):
        records = []

        for label_cell, row_number in label_cells:
            (
                label_col,
                target_col,
                ytd_col,
                achievement_col,
                current_col,
                required_col,
                estimated_col,
                projected_col,
            ) = start_columns

            records.append(
                {
                    "Segment": value(label_cell),
                    "Target": value(
                        f"{target_col}{row_number}"
                    ),
                    "YTD": value(
                        f"{ytd_col}{row_number}"
                    ),
                    "Achievement %": value(
                        f"{achievement_col}{row_number}"
                    ),                    "Current Monthly RR": value(
                        f"{current_col}{row_number}"
                    ),
                    "Required Monthly RR": value(
                        f"{required_col}{row_number}"
                    ),
                    "Estimated FY": value(
                        f"{estimated_col}{row_number}"
                    ),
                    "Projected Achievement %": value(
                        f"{projected_col}{row_number}"
                    ),
                }
            )

        return pd.DataFrame(records)

    asset_rows = [
        ("W8", 8),
        ("W10", 10),
        ("W11", 11),
        ("W12", 12),
    ]

    vertical_rows = [
        ("W14", 14),
        ("W15", 15),
        ("W16", 16),
        ("W17", 17),
        ("W18", 18),
    ]

    result["net_sales_asset"] = run_rate_table(
        asset_rows,
        (
            "W", "X", "Y", "Z",
            "AA", "AB", "AC", "AD",
        ),
    )

    result["net_sales_vertical"] = run_rate_table(
        vertical_rows,
        (
            "W", "X", "Y", "Z",
            "AA", "AB", "AC", "AD",
        ),
    )

    gs_asset_rows = [
        ("AF8", 8),
        ("AF10", 10),
        ("AF11", 11),
        ("AF12", 12),
    ]

    gs_vertical_rows = [
        ("AF14", 14),
        ("AF15", 15),
        ("AF16", 16),
        ("AF17", 17),
        ("AF18", 18),
    ]

    result["gross_sales_asset"] = run_rate_table(
        gs_asset_rows,
        (
            "AF", "AG", "AH", "AI",
            "AJ", "AK", "AL", "AM",
        ),
    )

    result["gross_sales_vertical"] = run_rate_table(
        gs_vertical_rows,
        (
            "AF", "AG", "AH", "AI",
            "AJ", "AK", "AL", "AM",
        ),
    )

    location_rows = [
        8, 10, 11, 12, 14, 15
    ]

    result["location_net_sales"] = pd.DataFrame(
        [
            {
                "Market Type": value(f"AQ{row}"),
                "Target": value(f"AR{row}"),
                "YTD": value(f"AS{row}"),
                "Achievement %": value(f"AT{row}"),
                "Current Monthly RR": value(
                    f"AU{row}"
                ),
                "Monthly Target RR": value(
                    f"AV{row}"
                ),
                "Estimated FY": value(f"AW{row}"),
                "Projected Achievement %": value(
                    f"AX{row}"
                ),
            }
            for row in location_rows
        ]
    )

    result["location_gross_sales"] = pd.DataFrame(
        [
            {
                "Market Type": value(f"BJ{row}"),
                "Target": value(f"BK{row}"),
                "YTD": value(f"BL{row}"),
                "Achievement %": value(f"BM{row}"),
                "Current Monthly RR": value(
                    f"BN{row}"
                ),
                "Monthly Target RR": value(
                    f"BO{row}"
                ),
                "Estimated FY": value(f"BP{row}"),
                "Projected Achievement %": value(
                    f"BQ{row}"
                ),
            }
            for row in location_rows
        ]
    )

    market_rows = list(range(7, 14))

    result["market_asset_cut"] = pd.DataFrame(
        [
            {
                "Market Type": value(f"CC{row}"),
                "Overall Target": value(f"CD{row}"),
                "Overall YTD": value(f"CE{row}"),
                "Overall Ach %": value(f"CF{row}"),
                "Equity Target": value(f"CG{row}"),
                "Equity YTD": value(f"CH{row}"),
                "Equity Ach %": value(f"CI{row}"),
                "Debt Target": value(f"CJ{row}"),
                "Debt YTD": value(f"CK{row}"),
                "Debt Ach %": value(f"CL{row}"),
                "Liquid Target": value(f"CM{row}"),
                "Liquid YTD": value(f"CN{row}"),
                "Liquid Ach %": value(f"CO{row}"),
            }
            for row in market_rows
            if value(f"CC{row}") is not None
        ]
    )

    result["budget_achievement"] = rows(
        [
            {
                "Asset Class": "CR9",
                "FY26 Budget": "CS9",
                "FY26 Retail": "CT9",
                "FY26 Insti": "CU9",
                "FY26 Actual": "CV9",
                "FY25 Budget": "CY9",
                "FY25 Retail": "CZ9",
                "FY25 Insti": "DA9",
                "FY25 Actual": "DB9",
            },
            {
                "Asset Class": "CR10",
                "FY26 Budget": "CS10",
                "FY26 Retail": "CT10",
                "FY26 Insti": "CU10",
                "FY26 Actual": "CV10",
                "FY25 Budget": "CY10",
                "FY25 Retail": "CZ10",
                "FY25 Insti": "DA10",
                "FY25 Actual": "DB10",
            },
            {
                "Asset Class": "CR11",
                "FY26 Budget": "CS11",
                "FY26 Retail": "CT11",
                "FY26 Insti": "CU11",
                "FY26 Actual": "CV11",
                "FY25 Budget": "CY11",
                "FY25 Retail": "CZ11",
                "FY25 Insti": "DA11",
                "FY25 Actual": "DB11",
            },
            {
                "Asset Class": "CR12",
                "FY26 Budget": "CS12",
                "FY26 Retail": "CT12",
                "FY26 Insti": "CU12",
                "FY26 Actual": "CV12",
                "FY25 Budget": "CY12",
                "FY25 Retail": "CZ12",
                "FY25 Insti": "DA12",
                "FY25 Actual": "DB12",
            },
        ]
    )

    result["projection_comparison"] = rows(
        [
            {
                "Asset Class": "DL8",
                "Current RR Projection %": "DM8",
                "Simulation %": "DS8",
                "Changed Projection": "DY8",
                "Current RR Number": "EE8",
                "FY27 Budget": "EK8",
            },
            {
                "Asset Class": "DL10",
                "Current RR Projection %": "DN8",
                "Simulation %": "DT8",
                "Changed Projection": "DZ8",
                "Current RR Number": "EF8",
                "FY27 Budget": "EL8",
            },
            {
                "Asset Class": "DL11",
                "Current RR Projection %": "DO8",
                "Simulation %": "DU8",
                "Changed Projection": "EA8",
                "Current RR Number": "EG8",
                "FY27 Budget": "EM8",
            },
            {
                "Asset Class": "DL12",
                "Current RR Projection %": "DP8",
                "Simulation %": "DV8",
                "Changed Projection": "EB8",
                "Current RR Number": "EH8",
                "FY27 Budget": "EN8",
            },
        ]
    )

    result["projection_comparison"].iloc[:, 0] = [
        "Overall",
        "Equity",
        "Debt",
        "Liquid",
    ]

    result["achievement_buckets"] = pd.DataFrame(
        [
            {
                "Achievement Band": value(
                    f"FN{row}"
                ),
                "Current": value(f"FO{row}"),
                "+5% RR": value(f"FP{row}"),
                "+10% RR": value(f"FQ{row}"),
                "+15% RR": value(f"FR{row}"),
            }
            for row in range(8, 14)
        ]
    )

    result["scenario_revenue"] = rows(
        [
            {
                "Scenario": "FO15",
                "Projected Equity NS": "FO16",
                "Achievement %": "FO17",
                "Revenue/PBT": "FO18",
            },
            {
                "Scenario": "FP6",
                "Projected Equity NS": "FP16",
                "Achievement %": "FP17",
                "Revenue/PBT": "FP18",
            },
            {
                "Scenario": "FQ6",
                "Projected Equity NS": "FQ16",
                "Achievement %": "FQ17",
                "Revenue/PBT": "FQ18",
            },
            {
                "Scenario": "FR6",
                "Projected Equity NS": "FR16",
                "Achievement %": "FR17",
                "Revenue/PBT": "FR18",
            },
        ]
    )

    result["scenario_revenue"].iloc[:, 0] = [
        "Current RR",
        "+5% RR",
        "+10% RR",
        "+15% RR",
    ]

    result["contest"] = pd.DataFrame(
        [
            {
                "Market Type": value(f"GB{row}"),
                "Target Distribution %": value(
                    f"GC{row}"
                ),
                "Total RMs": value(f"GD{row}"),
                "Selected RMs": value(f"GE{row}"),
                "Gate Achievement": value(
                    f"GF{row}"
                ),
                "Destination": value(f"GG{row}"),
                "Cost ₹": value(f"GH{row}"),
                "Already Qualified": value(
                    f"GI{row}"
                ),
            }
            for row in range(8, 13)
        ]
    )

    result["target_buckets"] = pd.DataFrame(
        [
            {
                "Target Cap": value(f"GK{row}"),
                "#RMs": value(f"GL{row}"),
                "Equity Target": value(f"GM{row}"),
            }
            for row in range(7, 16)
        ]
    )

    result["option_two"] = rows(
        [
            {
                "Target Group": "GO7",
                "#RMs": "GP7",
                "Equity Target": "GQ7",
                "Projected Current Achievement": "GR7",
                "Currently Qualifying": "GS7",
            },
            {
                "Target Group": "GO8",
                "#RMs": "GP8",
                "Equity Target": "GQ8",
                "Projected Current Achievement": "GR8",
                "Currently Qualifying": "GS8",
            },
            {
                "Target Group": "GO9",
                "#RMs": "GP9",
                "Equity Target": "GQ9",
                "Projected Current Achievement": "GR9",
                "Currently Qualifying": "GS9",
            },
        ]
    )

    result["scenario_two"] = rows(
        [
            {
                "Asset Class": "S59",
                "FY27 Target": "T59",
                "Estimated Achievement": "U59",
                "Required RR": "W59",
                "Current RR": "X59",
                "Additional RR": "Y59",
                "Revenue": "Z59",
            },
            {
                "Asset Class": "S60",
                "FY27 Target": "T60",
                "Estimated Achievement": "U60",
                "Required RR": "W60",
                "Current RR": "X60",
                "Additional RR": "Y60",
                "Revenue": "Z60",
            },
            {
                "Asset Class": "S61",
                "FY27 Target": "T61",
                "Estimated Achievement": "U61",
                "Required RR": "W61",
                "Current RR": "X61",
                "Additional RR": "Y61",
                "Revenue": "Z61",
            },
        ]
    )

    result["scenario_four"] = rows(
        [
            {
                "Asset Class": "AD59",
                "FY27 Target": "AE59",
                "Estimated Achievement": "AF59",
                "Estimated Achievement %": "AG59",
                "Required RR": "AH59",
                "Current RR": "AI59",
                "RR After Leakage": "AJ59",
                "Estimated FY": "AK59",
                "Revenue": "AL59",
            },
            {
                "Asset Class": "AD60",
                "FY27 Target": "AE60",
                "Estimated Achievement": "AF60",
                "Estimated Achievement %": "AG60",
                "Required RR": "AH60",
                "Current RR": "AI60",
                "RR After Leakage": "AJ60",
                "Estimated FY": "AK60",
                "Revenue": "AL60",
            },
            {
                "Asset Class": "AD61",
                "FY27 Target": "AE61",
                "Estimated Achievement": "AF61",
                "Estimated Achievement %": "AG61",
                "Required RR": "AH61",
                "Current RR": "AI61",
                "RR After Leakage": "AJ61",
                "Estimated FY": "AK61",
                "Revenue": "AL61",
            },
        ]
    )

    result["months_done"] = value("W6")
    result["employee_count"] = value("EQ6")

    result["formula_count"] = sum(
        1
        for row in formula_sheet.iter_rows()
        for cell in row
        if (
            isinstance(cell.value, str)
            and cell.value.startswith("=")
        )
        or hasattr(cell.value, "text")
    )

    result["audit_formulas"] = {
        "Gross Sales achievement": formula("E6"),
        "Net Sales current run rate": formula("AA8"),
        "Net Sales required run rate": formula("AB8"),
        "Net Sales full-year estimate": formula("AC8"),
        "Net Sales projected achievement": formula("AD8"),
        "Location-wise aggregation": formula("CG7"),
        "Scenario revenue": formula("FP18"),
        "Contest cost": formula("GH8"),
        "Qualification count": formula("GI8"),
    }

    return result


def final_sheet_page():
    st.html(
        """
<div class="hero">
    <div class="eyebrow">
        Page 2 · Authoritative Excel Model
    </div>

    <div class="hero-title">
        FINAL Sheet Analysis
    </div>

    <div class="hero-sub">
        Every section below is read from the cached outputs
        of the workbook's FINAL sheet. The layout converts
        the 209-column Excel dashboard into a
        management-ready calculation story.
    </div>
</div>
"""
    )

    with st.sidebar:
        st.divider()
        st.markdown("### FINAL Workbook")

        final_upload = st.file_uploader(
            "Upload Target Projections Workbook",
            type=["xlsx"],
            key="final_workbook_upload",
        )

    if final_upload is None:
        st.info(
            "Upload the workbook containing the "
            "FINAL sheet from the sidebar."
        )
        return

    try:
        model = load_final_analysis(
            final_upload.getvalue()
        )
    except Exception as error:
        st.error(
            f"Could not read FINAL sheet: {error}"
        )
        return

    current_ns = model["net_sales_asset"].iloc[0]
    current_gs = model["gross_sales_asset"].iloc[0]

    cards = st.columns(5)

    metrics = [
        (
            "Excel Formulas Read",
            f"{model['formula_count']:,}",
            "From FINAL sheet",
        ),
        (
            "Months Completed",
            int(model["months_done"]),
            "FINAL!W6 assumption",
        ),
        (
            "Employees",
            int(model["employee_count"]),
            "Overall employee base",
        ),
        (
            "Projected NS Achievement",
            pct(
                current_ns[
                    "Projected Achievement %"
                ] * 100
            ),
            fmt(current_ns["Estimated FY"]),
        ),
        (
            "Projected GS Achievement",
            pct(
                current_gs[
                    "Projected Achievement %"
                ] * 100
            ),
            fmt(current_gs["Estimated FY"]),
        ),
    ]

    for column, metric in zip(cards, metrics):
        with column:
            kpi(*metric)

    st.html(
        """
<div class="callout">
    <b style="color:#D4AF37">
        How to read this page
    </b>

    <br><br>

    Target is the full-year FY27 target.
    YTD is achievement after the completed months.
    Current RR is YTD ÷ months completed.
    Estimated FY is Current RR × 12.
    Projected achievement is Estimated FY ÷ Target.
    Negative achievement is displayed as a dash
    in several original Excel blocks.
</div>
"""
    )

    section(
        "1. Gross Sales Target Achievement",
        "Exact FINAL!B5:E9 asset-class totals.",
    )
    show_table(model["gross_sales_summary"])

    section(
        "2. Net Sales Run-Rate Analysis",
        "Asset-class and vertical projections using "
        "the FINAL sheet's four-month assumption.",
    )

    ns_asset_tab, ns_vertical_tab = st.tabs(
        ["Asset Class", "Vertical"]
    )

    with ns_asset_tab:
        show_table(model["net_sales_asset"])

    with ns_vertical_tab:
        show_table(model["net_sales_vertical"])

    section(
        "3. Gross Sales Run-Rate Analysis",
        "Current RR, full-year estimate and projected "
        "achievement for GS.",
    )

    gs_asset_tab, gs_vertical_tab = st.tabs(
        ["Asset Class", "Vertical"]
    )

    with gs_asset_tab:
        show_table(model["gross_sales_asset"])

    with gs_vertical_tab:
        show_table(model["gross_sales_vertical"])

    section(
        "4. Location-Wise Targets",
        "Market-type cuts for Overall, T2, T6, "
        "T30, B30 and EM.",
    )

    location_ns_tab, location_gs_tab = st.tabs(
        ["Net Sales", "Gross Sales"]
    )

    with location_ns_tab:
        show_table(model["location_net_sales"])

    with location_gs_tab:
        show_table(model["location_gross_sales"])

    section(
        "5. Market Type × Asset Class",
        "SUMIFS-based target, YTD and achievement "
        "calculations from RM Retail Sales.",
    )

    show_table(
        model["market_asset_cut"],
        430,
    )

    section(
        "6. Budget vs Achievement",
        "FY26 and FY25 budget/actual comparison split "
        "between Retail and Institutional business.",
    )

    show_table(model["budget_achievement"])

    section(
        "7. Projection Bridge",
        "Current-run-rate projection, simulation, "
        "changed numbers, current RR numbers and "
        "FY27 budget.",
    )

    show_table(model["projection_comparison"])

    section(
        "8. Equity Achievement Bucketing",
        "Number of RMs moving between achievement "
        "bands under current, +5%, +10% and +15% "
        "run rates.",
    )

    bucket_left, bucket_right = st.columns(
        [1.25, 1],
        gap="large",
    )

    with bucket_left:
        show_table(model["achievement_buckets"])

    with bucket_right:
        fig = go.Figure()

        for column, color in [
            ("Current", "#666666"),
            ("+5% RR", "#9A7D24"),
            ("+10% RR", GOLD),
            ("+15% RR", TEXT),
        ]:
            fig.add_bar(
                x=model["achievement_buckets"][
                    "Achievement Band"
                ],
                y=model["achievement_buckets"][column],
                name=column,
                marker_color=color,
            )

        fig.update_layout(
            title=(
                "RM Movement Across "
                "Achievement Bands"
            ),
            barmode="group",
            xaxis_title="",
            yaxis_title="# RMs",
        )

        st.plotly_chart(
            chart_style(fig, 390),
            width="stretch",
            config={"displayModeBar": False},
        )

    section(
        "9. Scenario NS and Revenue",
        "Revenue/PBT is calculated in FINAL as "
        "projected Equity NS × 0.6%.",
    )

    show_table(model["scenario_revenue"])

    section(
        "10. Scenario 2 and Scenario 4",
        "Scenario 2 applies asset-class achievement "
        "assumptions. Scenario 4 additionally applies "
        "leakage to the required run rate.",
    )

    scenario_two_tab, scenario_four_tab = st.tabs(
        [
            "Scenario 2 · No Leakage",
            "Scenario 4 · With Leakage",
        ]
    )

    with scenario_two_tab:
        show_table(model["scenario_two"])

    with scenario_four_tab:
        show_table(model["scenario_four"])

    section(
        "11. Contest Economics",
        "Selected RMs × destination cost; "
        "qualification is counted using the "
        "market-specific achievement gate.",
    )
    show_table(model["contest"])
    section(
        "12. Target Bucketing Options",
        "Option 1 uses ₹5 Cr target caps; "
        "Option 2 divides RMs below and above ₹6.5 Cr.",
    )

    option_one_tab, option_two_tab = st.tabs(
        [
            "Option 1 · Target Caps",
            "Option 2 · ₹6.5 Cr Split",
        ]
    )

    with option_one_tab:
        show_table(model["target_buckets"])

    with option_two_tab:
        show_table(model["option_two"])

    section(
        "13. Calculation Dictionary",
        "Plain-English meaning of the formulas "
        "used repeatedly across FINAL.",
    )

    calculation_dictionary = pd.DataFrame(
        [
            [
                "Achievement %",
                "YTD ÷ Target",
                "Shows the proportion of the annual "
                "target already achieved.",
            ],
            [
                "Current Monthly RR",
                "YTD ÷ Months Completed",
                "Average achievement per completed month.",
            ],
            [
                "Required Monthly RR",
                "(Target − YTD) ÷ Remaining Months",
                "Monthly performance needed to close "
                "the remaining target gap.",
            ],
            [
                "Estimated FY",
                "Current Monthly RR × 12",
                "Annualises the current performance pace.",
            ],
            [
                "Projected Achievement %",
                "Estimated FY ÷ Target",
                "Expected percentage of target achieved "
                "if the current pace continues.",
            ],
            [
                "Market-Type Target/YTD",
                "SUMIFS by MKT TYPE",
                "Aggregates RM Retail Sales for T2, T6, "
                "T30, B30, EM and other market types.",
            ],
            [
                "New Qualifiers",
                "COUNTIFS above achievement gate",
                "Counts RMs whose Equity achievement is "
                "above the market-specific gate.",
            ],
            [
                "Contest Cost",
                "Selected RMs × Cost per Destination",
                "Japan uses ₹3.5 lakh and Turkey uses "
                "₹1.5 lakh per selected RM.",
            ],
            [
                "Revenue/PBT",
                "Projected Equity NS × 0.60 ÷ 100",
                "Converts projected Equity Net Sales "
                "to revenue at 0.6%.",
            ],
            [
                "Leakage-adjusted RR",
                "Required RR × (1 − Leakage %)",
                "Reduces the required run rate by the "
                "leakage assumption in Scenario 4.",
            ],
        ],
        columns=[
            "Calculation",
            "Formula Logic",
            "Explanation",
        ],
    )

    show_table(calculation_dictionary)

    with st.expander(
        "Formula audit — exact Excel formulas"
    ):
        audit = pd.DataFrame(
            [
                {
                    "Calculation": name,
                    "Excel Formula": formula_text,
                }
                for name, formula_text
                in model["audit_formulas"].items()
            ]
        )

        show_table(audit)

    st.html(
        """
<div class="callout">
    <b style="color:#D4AF37">
        Audit note
    </b>

    <br><br>

    The FINAL sheet uses different meanings for some
    columns carrying the same label.

    In the overall Net Sales block, Required RR is
    (Target − YTD) ÷ remaining months.

    In several location-wise blocks, the similarly
    labelled column is Target ÷ 12, which is the
    monthly target pace.

    This page preserves the workbook's outputs and
    names the latter “Monthly Target RR” to avoid
    mixing the two concepts.
</div>
"""
    )


if APP_PAGE == "2 · FINAL Sheet Analysis":
    final_sheet_page()
    st.stop()


st.html(
    """
<div class="hero">
    <div class="eyebrow">
        Executive Performance Review
    </div>

    <div class="hero-title">
        Run Rate Analysis
    </div>

    <div class="hero-sub">
        Upload the RM workbook to compare the current
        monthly run rate with +5%, +10%, +15% and a
        custom uplift. All results respond to the
        sidebar filters.
    </div>
</div>
"""
)


with st.sidebar:
    st.markdown("### Upload Data")

    uploaded = st.file_uploader(
        "Upload RM Workbook",
        type=["xlsx"],
    )


if uploaded is None:
    st.info(
        "Upload the RM Excel workbook "
        "from the sidebar to begin."
    )
    st.stop()


try:
    data = load_workbook(uploaded.getvalue())
except Exception as error:
    st.error(f"Could not read workbook: {error}")
    st.stop()


valid_identity = (
    data["Employee Name"].ne("")
    | data["Emp Code"].ne("")
)

valid_numeric = (
    data["FY 26 TGT EQ NS"].notna()
    & data["Equity NS Ach YTD June"].notna()
)

data = data[
    valid_identity
    & valid_numeric
    & (data["FY 26 TGT EQ NS"] > 0)
].copy()


with st.sidebar:
    st.divider()
    st.markdown("### Global Filters")

    market_options = sorted(
        data["Market Type"]
        .dropna()
        .unique()
        .tolist()
    )

    selected_market = st.selectbox(
        "MKT TYPE",
        ["All"] + market_options,
    )

    if selected_market != "All":
        data = data[
            data["Market Type"] == selected_market
        ].copy()

    filter_definitions = [
        ("Status", "Status"),
        ("Type", "Type"),
        ("ZONE", "Zone"),
        ("REGION", "Region"),
    ]

    for column, label in filter_definitions:
        if column not in data.columns:
            continue

        options = sorted(
            data[column]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )

        if (
            column == "Status"
            and "Active" in options
        ):
            default = ["Active"]
        else:
            default = options

        selected = st.multiselect(
            label,
            options,
            default=default,
        )

        data = data[
            data[column].isin(selected)
        ].copy()

    st.divider()

    months = st.select_slider(
        "Projection Months After June",
        options=list(MONTHS),
        value=9,
        format_func=lambda value: (
            f"{value} months · "
            f"through {MONTHS[value]}"
        ),
    )

    custom_uplift = st.number_input(
        "Custom Run Rate Increase (%)",
        min_value=0.0,
        max_value=500.0,
        value=20.0,
        step=1.0,
    )


if data.empty:
    st.warning(
        "No valid RMs remain after applying "
        "the selected filters."
    )
    st.stop()


specs = [
    ("Current", 0.0),
    ("+5%", 5.0),
    ("+10%", 10.0),
    ("+15%", 15.0),
    (
        f"Custom +{custom_uplift:.1f}%",
        custom_uplift,
    ),
]

frames = {
    label: project(data, months, uplift)
    for label, uplift in specs
}

baseline = frames["Current"]

comparison = pd.DataFrame(
    [
        summarize(
            frames[label],
            label,
            uplift,
            baseline,
        )
        for label, uplift in specs
    ]
)

current = comparison.iloc[0]
custom = comparison.iloc[-1]

custom_frame = frames[
    f"Custom +{custom_uplift:.1f}%"
]

target_gap = max(
    current["Target"] - current["Projected NS"],
    0,
)

gap_closed = max(
    custom["Projected NS"] - current["Projected NS"],
    0,
)

gap_closed_pct = (
    min(gap_closed / target_gap * 100, 100)
    if target_gap else 100
)

extra_rms = int(
    custom["RMs ≥100%"]
    - current["RMs ≥100%"]
)


st.html(
    f"""
<div class="callout">
    <b style="color:#D4AF37">Scope:</b>

    {selected_market} market ·
    {len(data):,} valid RMs ·
    projection through {MONTHS[months]}.

    Revenue is calculated as
    <b style="color:#F3F0E7">
        Net Sales × 0.60 / 100
    </b>.
</div>
"""
)


columns = st.columns(5)

cards = [
    (
        "Total Target",
        fmt(current["Target"]),
        f"{len(data):,} RMs in scope",
    ),
    (
        "Current Achievement",
        pct(current["Portfolio Achievement %"]),
        f"Projected NS {fmt(current['Projected NS'])}",
    ),
    (
        "Custom Achievement",
        pct(custom["Portfolio Achievement %"]),
        f"At +{custom_uplift:.1f}% run rate",
    ),
    (
        "Additional RMs ≥100%",
        f"+{extra_rms:,}",
        (
            f"{int(current['RMs ≥100%']):,} → "
            f"{int(custom['RMs ≥100%']):,}"
        ),
    ),
    (
        "Incremental Net Sales",
        fmt(custom["Incremental NS"]),
        (
            "Revenue impact "
            f"{fmt(custom['Incremental Revenue'])}"
        ),
    ),
]

for column, card in zip(columns, cards):
    with column:
        kpi(*card)


if custom["Portfolio Achievement %"] >= 100:
    conclusion = (
        "The custom scenario takes the filtered "
        "portfolio above its total target."
    )
elif gap_closed_pct >= 50:
    conclusion = (
        "The custom uplift closes a meaningful "
        "portion of the target gap, but additional "
        "action is still required."
    )
else:
    conclusion = (
        "Run-rate uplift alone is insufficient; "
        "the closest-to-target RMs should become "
        "the immediate action pool."
    )


st.html(
    f"""
<div class="callout">
    <b style="color:#D4AF37">
        Executive conclusion
    </b>

    <br><br>

    {conclusion}

    A +{custom_uplift:.1f}% run-rate increase adds
    <b>{fmt(custom['Incremental NS'])}</b>
    in projected NS, creates
    <b>{extra_rms:,} new qualifiers</b>,
    and closes
    <b>{pct(gap_closed_pct)}</b>
    of the current target gap.
</div>
"""
)


section(
    "1. Scenario Scorecard",
    "Management comparison of achievement, "
    "qualification, Net Sales and Revenue.",
)

show_table(
    comparison[
        [
            "Scenario",
            "RR Uplift %",
            "Projected NS",
            "Portfolio Achievement %",
            "RMs ≥100%",
            "Qualification Rate %",
            "New Qualifiers",
            "Incremental NS",
            "Revenue",
            "Incremental Revenue",
        ]
    ]
)


section(
    "2. Portfolio Movement",
    "The left chart shows portfolio delivery; "
    "the right chart shows people conversion.",
)

left, right = st.columns(2, gap="large")

with left:
    fig = go.Figure(
        go.Bar(
            x=comparison["Scenario"],
            y=comparison[
                "Portfolio Achievement %"
            ],
            marker_color=(
                ["#555555"]
                + [GOLD] * (len(comparison) - 1)
            ),
            text=[
                pct(value)
                for value in comparison[
                    "Portfolio Achievement %"
                ]
            ],
            textposition="outside",
        )
    )

    fig.add_hline(
        y=100,
        line_dash="dash",
        line_color=TEXT,
        annotation_text="100% target",
    )

    fig.update_layout(
        title="Portfolio Achievement by Scenario",
        xaxis_title="",
        yaxis_title="Achievement (%)",
    )

    st.plotly_chart(
        chart_style(fig),
        width="stretch",
        config={"displayModeBar": False},
    )


with right:
    fig = go.Figure(
        go.Scatter(
            x=comparison["Scenario"],
            y=comparison["RMs ≥100%"],
            mode="lines+markers+text",
            text=(
                comparison["RMs ≥100%"]
                .astype(int)
            ),
            textposition="top center",
            line=dict(
                color=GOLD,
                width=3,
            ),
            marker=dict(
                color=GOLD,
                size=10,
            ),
        )
    )

    fig.update_layout(
        title="RMs Crossing 100%",
        xaxis_title="",
        yaxis_title="Number of RMs",
    )

    st.plotly_chart(
        chart_style(fig),
        width="stretch",
        config={"displayModeBar": False},
    )


section(
    "3. Conversion Opportunity",
    "RMs nearest to 100% offer the fastest "
    "qualification opportunity.",
)

opportunity = baseline.copy()

opportunity["Achievement Band"] = pd.cut(
    opportunity["Projected Achievement %"],
    bins=[
        0,
        80,
        90,
        95,
        100,
        np.inf,
    ],
    labels=[
        "Below 80%",
        "80–90%",
        "90–95%",
        "95–100%",
        "100%+",
    ],
    right=False,
)

band_table = (
    opportunity
    .groupby(
        "Achievement Band",
        observed=False,
    )
    .agg(
        **{
            "#RMs": (
                "Employee Name",
                "size",
            ),
            "Target": (
                "FY 26 TGT EQ NS",
                "sum",
            ),
            "Projected NS": (
                "Projected Final NS",
                "sum",
            ),
        }
    )
    .reset_index()
)

band_table["Share of RMs %"] = (
    band_table["#RMs"] / len(data) * 100
)

show_table(band_table)


near_mask = baseline[
    "Projected Achievement %"
].between(
    90,
    100,
    inclusive="left",
)

near_count = int(near_mask.sum())

near_gap = (
    baseline.loc[
        near_mask,
        "FY 26 TGT EQ NS",
    ]
    - baseline.loc[
        near_mask,
        "Projected Final NS",
    ]
).clip(lower=0).sum()


section(
    "4. Market-Type Drivers",
    (
        f"Custom +{custom_uplift:.1f}% scenario "
        "contribution and qualification."
    ),
)

market_group = (
    custom_frame
    .groupby(
        "Market Type",
        dropna=False,
    )
    .agg(
        **{
            "#RMs": (
                "Employee Name",
                "size",
            ),
            "Target": (
                "FY 26 TGT EQ NS",
                "sum",
            ),
            "Projected NS": (
                "Projected Final NS",
                "sum",
            ),
            "RMs ≥100%": (
                "Qualifies",
                "sum",
            ),
        }
    )
    .reset_index()
)

market_group["Portfolio Achievement %"] = np.where(
    market_group["Target"] > 0,
    (
        market_group["Projected NS"]
        / market_group["Target"]
        * 100
    ),
    0,
)

market_group["Qualification Rate %"] = (
    market_group["RMs ≥100%"]
    / market_group["#RMs"]
    * 100
)

market_group["NS Contribution %"] = (
    market_group["Projected NS"]
    / market_group["Projected NS"].sum()
    * 100
)

new_market = (
    custom_frame.loc[
        custom_frame["Qualifies"]
        & ~baseline["Qualifies"]
    ]
    .groupby("Market Type")
    .size()
    .rename("New Qualifiers")
    .reset_index()
)

market_group = market_group.merge(
    new_market,
    on="Market Type",
    how="left",
)

market_group["New Qualifiers"] = (
    market_group["New Qualifiers"]
    .fillna(0)
    .astype(int)
)


chart_column, table_column = st.columns(
    [1, 1.2],
    gap="large",
)

with chart_column:
    plot_data = market_group.sort_values(
        "Projected NS"
    )

    fig = go.Figure(
        go.Bar(
            x=plot_data["Projected NS"],
            y=plot_data["Market Type"],
            orientation="h",
            marker_color=GOLD,
            text=plot_data[
                "Projected NS"
            ].map(fmt),
            textposition="auto",
        )
    )

    fig.update_layout(
        title="Projected NS by Market Type",
        xaxis_title="Projected NS",
        yaxis_title="",
    )

    st.plotly_chart(
        chart_style(fig, 420),
        width="stretch",
        config={"displayModeBar": False},
    )


with table_column:
    show_table(
        market_group.sort_values(
            "Projected NS",
            ascending=False,
        ),
        420,
    )


section(
    "5. Management Takeaways",
    "Ready-to-present summary for the "
    "performance review.",
)

left, right = st.columns(2, gap="large")

remaining_gap = max(
    custom["Target"] - custom["Projected NS"],
    0,
)


with left:
    st.html(
        f"""
<div class="management">
    <b>Portfolio story</b>

    <br><br>

    • Current achievement is
    {pct(current['Portfolio Achievement %'])}.

    <br>

    • +{custom_uplift:.1f}% run rate raises
    achievement to
    {pct(custom['Portfolio Achievement %'])}.

    <br>

    • Incremental NS is
    {fmt(custom['Incremental NS'])}.

    <br>

    • Incremental Revenue is
    {fmt(custom['Incremental Revenue'])}.

    <br>

    • Remaining target gap is
    {fmt(remaining_gap)}.
</div>
"""
    )


with right:
    st.html(
        f"""
<div class="management">
    <b>People story</b>

    <br><br>

    • Qualified RMs increase from
    {int(current['RMs ≥100%']):,}
    to
    {int(custom['RMs ≥100%']):,}.

    <br>

    • {extra_rms:,} RMs newly cross 100%.

    <br>

    • {near_count:,} RMs are currently in the
    90–100% conversion zone.

    <br>

    • Their combined gap to target is
    {fmt(near_gap)}.

    <br>

    • This group should receive the first
    focused intervention.
</div>
"""
    )


with st.expander("RM-Level Drill-Down"):
    detail = custom_frame.copy()

    detail["Current Achievement %"] = baseline[
        "Projected Achievement %"
    ]

    detail["Newly Qualifies"] = (
        detail["Qualifies"]
        & ~baseline["Qualifies"]
    )

    columns = [
        column
        for column in [
            "Emp Code",
            "Employee Name",
            "Market Type",
            "ZONE",
            "REGION",
            "FY 26 TGT EQ NS",
            "Equity NS Ach YTD June",
            "Current Monthly RR",
            "Current Achievement %",
            "Scenario Monthly RR",
            "Projected Final NS",
            "Projected Achievement %",
            "Newly Qualifies",
        ]
        if column in detail.columns
    ]

    show_table(
        detail[columns].sort_values(
            [
                "Newly Qualifies",
                "Projected Achievement %",
            ],
            ascending=[
                False,
                False,
            ],
        ),
        550,
    )
